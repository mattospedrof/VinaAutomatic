import os
import re
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas


menu = """
### Mattos Tech ###

~Analysis~

[A] I have a txt file of docking results and would like to get a spreadsheet and boxplot.
[B] I have an xlsx file with docking results and I would like to get the boxplot.
[C] Exit the system.
=> """

choose_menu = input(menu)

while True:
    match choose_menu.upper():
        case 'A':
            with open("output_results.txt", "r", encoding="utf-8") as file:
                lines = file.readlines()

            first_column_regex = r'(?:(?<=\n)|(?<=^))(\d+)\t(?=\d+\t-\d+\.\d+|\s*$)'
            second_column_regex = r'(?<=\t)(\d+)(?=\t)'
            third_column_regex = re.compile(r'\t\s*\d+\s+(-?\d+(?:\.\d+)?)\s*')

            results = []

            for line in lines:
                first_match = re.search(first_column_regex, line)
                second_match = re.search(second_column_regex, line)
                third_match = re.search(third_column_regex, line)
                
                results.append((int(first_match.group(1)) if first_match else '',
                                int(second_match.group(1)) if second_match else None,
                                float(third_match.group(1)) if third_match else None))
                
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Round', 'Mode', 'Affinity'])

            for row in results:
                sheet.append(row)

            for column in [1, 2]:
                col_letter = get_column_letter(column)
                for cell in sheet[col_letter][1:]:
                    cell.number_format = '0'

            workbook.save("output_results.xlsx")
            
            time.sleep(1)

            if os.path.exists("output_results.xlsx"):
                print("xlsx file created successfully.\n")
                print("...")
                wb = load_workbook(filename="output_results.xlsx")

                sheet = wb.active

                column_values = [cell.value for cell in sheet['C'][1:] if cell.value is not None]
                column_values = [value for value in column_values if value is not None and value != '']
                column_values = [0 if value is None else value for value in column_values]
                
                title_graph = input("Enter the desired title for the graph. If you don't want to, just hit enter.\n"
                                    "=> ")
                
                x_ticks = input("Enter the desired name for the ligand.\n"
                                "=> ")
                
                plt.rcParams.update({'font.family': 'serif', 'font.size': 12})
                plt.boxplot(column_values)
                plt.title(title_graph)
                plt.xlabel(' ')
                plt.ylabel(r'Affinity kcal$\cdot$mol$^{-1}$')
                plt.xticks([1], [x_ticks])
                plt.gca().spines['top'].set_visible(False)
                plt.gca().spines['right'].set_visible(False)

                print("...")
                canvas = FigureCanvas(plt.gcf())
                canvas.copy_from_bbox(canvas.figure.bbox)
                time.sleep(1)
                plt.show()

                save = input("Do you want save the figure? y/n ")

                match save.lower()[:1]:
                    case 'y':
                        canvas.print_figure('boxplot_results.png')
                        print("Verify your directory")
                        break
                    case 'n':
                        print("OK, thank you for using the script")
                        break
                    case _:
                        print("Please, enter a valid option.")
            else:
                print("Oops, something happened")
        case 'B':
            print("...")
            wb = load_workbook(filename="output_results.xlsx")

            sheet = wb.active

            column_values = [cell.value for cell in sheet['C'][1:] if cell.value is not None]
            column_values = [value for value in column_values if value is not None and value != '']
            column_values = [0 if value is None else value for value in column_values]
            
            title_graph = input("Enter the desired title for the graph. If you don't want to, just hit enter.\n"
                                "=> ")
            
            x_ticks = input("Enter the desired name for the ligand.\n"
                            "=> ")
                
            plt.rcParams.update({'font.family': 'serif', 'font.size': 12})
            plt.boxplot(column_values)
            plt.title(title_graph)
            plt.xlabel(' ')
            plt.ylabel("Afinidade (kcal.mol$^{-1}$)")
            plt.xticks([1], [x_ticks])
            plt.gca().spines['top'].set_visible(False)
            plt.gca().spines['right'].set_visible(False)

            print("...")
            canvas = FigureCanvas(plt.gcf())
            canvas.copy_from_bbox(canvas.figure.bbox)
            time.sleep(1)
            plt.show()

            save = input("Do you want save the figure? y/n ")

            match save.lower()[:1]:
                case 'y':
                    canvas.print_figure('boxplot_results.png')
                    print("Verify your directory")
                    break
                case 'n':
                    print("OK, thank you for using the script")
                    break
                case _:
                    print("Please, enter a valid option.")
        case 'C':
            print("Thank you for using our system")
            break
        case _:
            print("Please, enter a valid option.")
            break
